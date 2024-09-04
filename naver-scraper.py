import sys
import time
import logging
import os
import json
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QProgressBar, QSpinBox, QMessageBox, QMainWindow, QAction, QFileDialog, QTextEdit,
                             QGroupBox, QFormLayout, QStyleFactory, QListWidget, QDialog, QDialogButtonBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QIcon
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M')

def setup_driver():
    service = Service(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(service=service, options=options)

def scrape_restaurant_info(driver):
    try:
        driver.switch_to.default_content()
        WebDriverWait(driver, 90).until(
            EC.frame_to_be_available_and_switch_to_it((By.ID, 'entryIframe'))
        )
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'place_section'))
        )
        time.sleep(5)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        restaurant_name = soup.select_one('span.GHAhO')
        restaurant_name = restaurant_name.text.strip() if restaurant_name else "N/A"
        address = soup.select_one('span.LDgIH')
        address = address.text.strip() if address else "N/A"
        phone_number = soup.select_one('span.xlx7Q')
        phone_number = phone_number.text.strip() if phone_number else "N/A"
        return {
            "Restaurant Name": restaurant_name,
            "Address": address,
            "Phone Number": phone_number
        }
    except Exception as e:
        logging.error(f"An error occurred while scraping restaurant info: {e}")
        return None

def scrape_single_restaurant(driver, restaurant_index, max_retries=3):
    for attempt in range(max_retries):
        try:
            driver.switch_to.default_content()
            WebDriverWait(driver, 30).until(
                EC.frame_to_be_available_and_switch_to_it((By.ID, 'searchIframe'))
            )
            scroll_container = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, '_pcmap_list_scroll_container'))
            )
            scroll_amount = (restaurant_index - 1) * 100
            driver.execute_script(f"arguments[0].scrollTop += {scroll_amount};", scroll_container)
            time.sleep(2)
            selectors = [
                f'li.UEzoS.rTjJo:nth-child({restaurant_index}) .place_bluelink',
                f'div.lazyload-wrapper:nth-child({restaurant_index}) .place_bluelink'
            ]
            restaurant_element = None
            for selector in selectors:
                try:
                    restaurant_element = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    break
                except:
                    continue
            if not restaurant_element:
                raise NoSuchElementException(f"Could not find restaurant element for index {restaurant_index}")
            driver.execute_script("arguments[0].scrollIntoView();", restaurant_element)
            time.sleep(2)
            restaurant_element.click()
            time.sleep(5)
            return scrape_restaurant_info(driver)
        except Exception as e:
            logging.error(f"Attempt {attempt + 1} failed for restaurant {restaurant_index}: {e}")
            if attempt == max_retries - 1:
                logging.error(f"Failed to scrape restaurant {restaurant_index} after {max_retries} attempts")
                return None
        time.sleep(5)

class ScrapingTask:
    def __init__(self, address, search_query, num_restaurants, zoom_level):
        self.address = address
        self.search_query = search_query
        self.num_restaurants = num_restaurants
        self.zoom_level = zoom_level

    def to_dict(self):
        return {
            'address': self.address,
            'search_query': self.search_query,
            'num_restaurants': self.num_restaurants,
            'zoom_level': self.zoom_level
        }

    @classmethod
    def from_dict(cls, data):
        return cls(data['address'], data['search_query'], data['num_restaurants'], data['zoom_level'])

class AddTaskDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Scraping Task")
        self.layout = QVBoxLayout(self)

        self.address_input = QLineEdit(self)
        self.search_query_input = QLineEdit("근처 식당", self)
        self.num_restaurants_input = QSpinBox(self)
        self.zoom_level_input = QSpinBox(self)

        self.num_restaurants_input.setRange(1, 1000)
        self.num_restaurants_input.setValue(50)

        self.zoom_level_input.setRange(1, 100)
        self.zoom_level_input.setValue(50)

        form_layout = QFormLayout()
        form_layout.addRow("Address:", self.address_input)
        form_layout.addRow("Search Query:", self.search_query_input)
        form_layout.addRow("Number of Restaurants:", self.num_restaurants_input)
        form_layout.addRow("Zoom Level:", self.zoom_level_input)

        self.layout.addLayout(form_layout)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)

        self.layout.addWidget(self.button_box)

    def get_task(self):
        return ScrapingTask(
            self.address_input.text(),
            self.search_query_input.text(),
            self.num_restaurants_input.value(),
            self.zoom_level_input.value()
        )

class SchedulerWidget(QWidget):
    taskAdded = pyqtSignal(ScrapingTask)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.task_list = QListWidget()
        self.layout.addWidget(self.task_list)
        self.tasks = []

    def add_task(self):
        dialog = AddTaskDialog(self)
        if dialog.exec_():
            task = dialog.get_task()
            self.tasks.append(task)
            self.task_list.addItem(f"{task.address} - {task.search_query} - {task.num_restaurants} - {task.zoom_level}")
            self.taskAdded.emit(task)

    def remove_task(self):
        current_row = self.task_list.currentRow()
        if current_row >= 0:
            self.task_list.takeItem(current_row)
            del self.tasks[current_row]

    def get_tasks(self):
        return self.tasks

    def save_tasks(self, filename):
        with open(filename, 'w') as f:
            json.dump([task.to_dict() for task in self.tasks], f)

    def load_tasks(self, filename):
        with open(filename, 'r') as f:
            task_data = json.load(f)
            self.tasks = [ScrapingTask.from_dict(data) for data in task_data]
            self.task_list.clear()
            for task in self.tasks:
                self.task_list.addItem(f"{task.address} - {task.search_query} - {task.num_restaurants} - {task.zoom_level}")

class ScraperThread(QThread):
    progress_update = pyqtSignal(int)
    scraping_complete = pyqtSignal(str)
    data_scraped = pyqtSignal(dict)

    def __init__(self, address, search_query, num_restaurants, zoom_slider_value):
        QThread.__init__(self)
        self.address = address
        self.search_query = search_query
        self.num_restaurants = num_restaurants
        self.zoom_slider_value = zoom_slider_value
        self.stop_flag = False

    def run(self):
        driver = setup_driver()
        try:
            driver.get("https://map.naver.com/")
            search_input_box = WebDriverWait(driver, 90).until(
                EC.element_to_be_clickable((By.CLASS_NAME, 'input_search'))
            )
            search_input_box.send_keys(self.address)
            search_input_box.send_keys(Keys.RETURN)
            time.sleep(5)
            total_zoom_levels = 14
            desired_zoom_level = int((self.zoom_slider_value / 100) * (total_zoom_levels - 1))
            current_zoom_level = total_zoom_levels // 2
            zoom_difference = desired_zoom_level - current_zoom_level
            zoom_button_selector = 'button.zoom_in' if zoom_difference > 0 else 'button.zoom_out'
            zoom_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, zoom_button_selector))
            )
            for _ in range(abs(zoom_difference)):
                zoom_button.click()
                time.sleep(1)
            time.sleep(2)
            search_input_box = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.CLASS_NAME, 'input_search'))
            )
            search_input_box.clear()
            search_input_box.send_keys(Keys.CONTROL + "a")
            search_input_box.send_keys(Keys.DELETE)
            time.sleep(2)
            if search_input_box.get_attribute('value'):
                search_input_box.clear()
                search_input_box.send_keys(Keys.CONTROL + "a")
                search_input_box.send_keys(Keys.DELETE)
                time.sleep(1)
            search_input_box.send_keys(self.search_query)
            search_input_box.send_keys(Keys.RETURN)
            time.sleep(10)
            for i in range(1, self.num_restaurants + 1):
                if self.stop_flag:
                    self.scraping_complete.emit("Scraping stopped by user")
                    break
                restaurant_info = scrape_single_restaurant(driver, i)
                if restaurant_info:
                    self.data_scraped.emit(restaurant_info)
                time.sleep(5)
                progress = int((i / self.num_restaurants) * 100)
                self.progress_update.emit(progress)
            if not self.stop_flag:
                self.scraping_complete.emit("Scraping completed successfully!")
        except Exception as e:
            self.scraping_complete.emit(f"An error occurred: {str(e)}")
        finally:
            driver.quit()

    def stop(self):
        self.stop_flag = True

class LogHandler(logging.Handler):
    def __init__(self, signal):
        super().__init__()
        self.signal = signal

    def emit(self, record):
        log_entry = self.format(record)
        self.signal.emit(log_entry)

class ScraperGUI(QMainWindow):
    log_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.scraped_data = {}
        self.save_location = os.getcwd()
        self.excel_filename = "naver_restaurants_data.xlsx"
        self.current_task_index = 0
        self.initUI()
        self.setup_logging()
        self.scheduler_widget.task_list.itemDoubleClicked.connect(self.show_task_details)

    def setup_logging(self):
        self.log_handler = LogHandler(self.log_signal)
        self.log_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s',
                                                        datefmt='%Y-%m-%d %H:%M'))
        logging.getLogger().addHandler(self.log_handler)
        logging.getLogger().setLevel(logging.INFO)
        self.log_signal.connect(self.update_log_window)

    def update_log_window(self, log_entry):
        self.log_window.append(log_entry)

    def initUI(self):
        self.setWindowTitle('Naver Map Scraper')
        self.setMinimumSize(550, 750)

        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Task Management Section
        self.scheduler_widget = SchedulerWidget()

        task_group = QGroupBox("Task Management")
        task_layout = QVBoxLayout(task_group)

        task_button_layout = QHBoxLayout()
        self.add_task_button = QPushButton('Add Task')
        self.add_task_button.clicked.connect(self.scheduler_widget.add_task)
        task_button_layout.addWidget(self.add_task_button)

        self.remove_task_button = QPushButton('Remove Task')
        self.remove_task_button.clicked.connect(self.scheduler_widget.remove_task)
        task_button_layout.addWidget(self.remove_task_button)

        task_layout.addLayout(task_button_layout)
        task_layout.addWidget(self.scheduler_widget)
        main_layout.addWidget(task_group)

        # Control Buttons
        control_group = QGroupBox("Controls")
        control_layout = QHBoxLayout(control_group)

        self.start_button = QPushButton('START')
        self.start_button.clicked.connect(self.start_scheduled_scraping)
        control_layout.addWidget(self.start_button)

        self.stop_button = QPushButton('STOP')
        self.stop_button.clicked.connect(self.stop_scraping)
        self.stop_button.setEnabled(False)
        control_layout.addWidget(self.stop_button)

        main_layout.addWidget(control_group)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(self.progress_bar)

        # Log Window
        log_group = QGroupBox("Logs")
        log_layout = QVBoxLayout(log_group)
        self.toggle_log_button = QPushButton("Hide Logs")
        self.toggle_log_button.clicked.connect(self.toggle_log_visibility)
        log_layout.addWidget(self.toggle_log_button)

        self.log_window = QTextEdit(self)
        self.log_window.setReadOnly(True)
        log_layout.addWidget(self.log_window)
        main_layout.addWidget(log_group)

        # Menu Bar
        self.create_menu_bar()

        # Set global stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f4f4f4;
            }
            QLabel {
                font-size: 14px;
                color: #333333;
            }
            QLineEdit, QSpinBox, QTextEdit {
                padding: 8px;
                border: 1px solid #cccccc;
                border-radius: 5px;
                background-color: #ffffff;
            }
            QPushButton {
                background-color: #2db400;
                color: white;
                border-radius: 5px;
                padding: 8px 15px;
            }
            QPushButton:hover {
                background-color: #259a00;
            }
            QPushButton:disabled {
                background-color: #a9a9a9;
            }
            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 5px;
                text-align: center;
                color: #333333;
            }
            QProgressBar::chunk {
                background-color: #2db400;
            }
            QListWidget {
                background-color: #ffffff;
                border: 1px solid #cccccc;
                border-radius: 5px;
            }
            QGroupBox {
                border: 1px solid #cccccc;
                border-radius: 5px;
                margin-top: 15px;
                padding: 15px;
                background-color: #f4f4f4;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #2db400;
            }
            QTextEdit {
                border: 1px solid #cccccc;
                border-radius: 5px;
                background-color: #ffffff;
                color: #333333;
                font-family: Courier, monospace;
                font-size: 12px;
            }
        """)

    def create_menu_bar(self):
        menubar = self.menuBar()
        
        file_menu = menubar.addMenu('File')
        change_location_action = QAction('Change Save Location', self)
        change_location_action.triggered.connect(self.change_save_location)
        file_menu.addAction(change_location_action)

        rename_file_action = QAction('Rename Output File', self)
        rename_file_action.triggered.connect(self.rename_file)
        file_menu.addAction(rename_file_action)

        file_menu.addSeparator()

        exit_action = QAction('Exit', self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        scheduler_menu = menubar.addMenu('Scheduler')
        save_tasks_action = QAction('Save Tasks', self)
        save_tasks_action.triggered.connect(self.save_tasks)
        scheduler_menu.addAction(save_tasks_action)

        load_tasks_action = QAction('Load Tasks', self)
        load_tasks_action.triggered.connect(self.load_tasks)
        scheduler_menu.addAction(load_tasks_action)

        help_menu = menubar.addMenu('Help')
        about_action = QAction('About', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def toggle_log_visibility(self):
        if self.log_window.isVisible():
            self.log_window.hide()
            self.toggle_log_button.setText("Show Logs")
        else:
            self.log_window.show()
            self.toggle_log_button.setText("Hide Logs")

    def change_save_location(self):
        new_location = QFileDialog.getExistingDirectory(self, "Select Save Location")
        if new_location:
            self.save_location = new_location
            self.show_info_message("Save Location", f"Save location changed to: {self.save_location}")

    def rename_file(self):
        new_name, ok = QFileDialog.getSaveFileName(self, "Rename Excel File", self.excel_filename, "Excel Files (*.xlsx)")
        if ok and new_name:
            self.excel_filename = os.path.basename(new_name)
            self.show_info_message("File Renamed", f"Excel file will be saved as: {self.excel_filename}")

    def save_tasks(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Save Tasks", "", "JSON Files (*.json)")
        if filename:
            self.scheduler_widget.save_tasks(filename)
            self.show_info_message("Tasks Saved", f"Tasks saved to: {filename}")

    def load_tasks(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Load Tasks", "", "JSON Files (*.json)")
        if filename:
            self.scheduler_widget.load_tasks(filename)
            self.show_info_message("Tasks Loaded", f"Tasks loaded from: {filename}")

    def show_about(self):
        about_text = """
        <h2>Naver Map Scraper</h2>
        <p>This program scrapes restaurant information from Naver Maps using Selenium and ChromeDriver.</p>
        <p><b>Author:</b> Jonathan Booker Nelson</p>
        <p><b>Version:</b> 1.2</p>
        <p><b>GitHub:</b> <a href="https://github.com/booknite/">https://github.com/booknite/</a></p>
        """
        QMessageBox.about(self, "About Naver Map Scraper", about_text)

    def show_info_message(self, title, message):
        QMessageBox.information(self, title, message)

    def show_error_message(self, title, message):
        QMessageBox.critical(self, title, message)

    def start_scheduled_scraping(self):
        tasks = self.scheduler_widget.get_tasks()
        if not tasks:
            self.show_error_message("No Tasks", "Please add at least one scraping task.")
            return

        self.current_task_index = 0
        self.start_next_task()

    def start_next_task(self):
        tasks = self.scheduler_widget.get_tasks()
        if self.current_task_index < len(tasks):
            task = tasks[self.current_task_index]
            self.start_scraping(task)
        else:
            self.show_info_message("Scraping Complete", "All scheduled tasks have been completed.")
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
            self.save_scraped_data()

    def start_scraping(self, task):
        self.current_task = task
        self.scraped_data[task.address] = []
        self.scraper_thread = ScraperThread(task.address, task.search_query, task.num_restaurants, task.zoom_level)
        self.scraper_thread.progress_update.connect(self.update_progress)
        self.scraper_thread.scraping_complete.connect(self.scraping_finished)
        self.scraper_thread.data_scraped.connect(self.add_scraped_data)
        self.scraper_thread.start()
        
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)

    def stop_scraping(self):
        if hasattr(self, 'scraper_thread'):
            self.scraper_thread.stop()
        self.stop_button.setEnabled(False)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def scraping_finished(self, message):
        self.progress_bar.setValue(100)
        logging.info(message)
        self.current_task_index += 1
        self.start_next_task()

    def add_scraped_data(self, data):
        self.scraped_data[self.current_task.address].append(data)
        logging.info(f"Scraped data for {data['Restaurant Name']}")

    def save_scraped_data(self):
        if self.scraped_data:
            file_path = os.path.join(self.save_location, self.excel_filename)
            workbook = openpyxl.Workbook()
            
            # Remove the default sheet created by openpyxl
            workbook.remove(workbook.active)
            
            for task_address, task_data in self.scraped_data.items():
                # Get the search query corresponding to the address
                task_query = next(task.search_query for task in self.scheduler_widget.get_tasks() if task.address == task_address)
                # Create a sheet title using address and search query
                sheet_title = f"{task_address} - {task_query}"

                # Excel sheet names are limited to 31 characters
                if len(sheet_title) > 31:
                    sheet_title = sheet_title[:28] + "..."

                # Create a new sheet with the generated title
                sheet = workbook.create_sheet(title=sheet_title)
                
                # Write headers
                headers = ["Restaurant Name", "Address", "Phone Number"]
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # Write data
                for row, restaurant in enumerate(task_data, start=2):
                    for col, key in enumerate(headers, start=1):
                        sheet.cell(row=row, column=col, value=restaurant.get(key, "N/A"))
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.show_info_message("Data Saved", f"All scraped data has been saved to {file_path}")
        else:
            self.show_error_message("No Data", "No data was scraped.")
        self.scraped_data = {}  # Clear the data after saving

    def closeEvent(self, event):
        if hasattr(self, 'scraper_thread') and self.scraper_thread.isRunning():
            reply = QMessageBox.question(self, 'Exit',
                                         "A scraping task is still running. Are you sure you want to quit?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.scraper_thread.stop()
                self.scraper_thread.wait()
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    def show_task_details(self, item):
        task = self.scheduler_widget.tasks[self.scheduler_widget.task_list.row(item)]
        details = f"""
        <b>Address:</b> {task.address}
        <b>Search Query:</b> {task.search_query}
        <b>Number of Restaurants:</b> {task.num_restaurants}
        <b>Zoom Level:</b> {task.zoom_level}
        """
        QMessageBox.information(self, "Task Details", details)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))  # Use Fusion style for a modern look
    ex = ScraperGUI()
    ex.show()
    sys.exit(app.exec_())

